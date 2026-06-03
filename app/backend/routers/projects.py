"""Projects, Contacts, and Graph API endpoints."""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import openpyxl
from bson import ObjectId
from fastapi import APIRouter, Body, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill

from core.database import (
    count_contacts,
    count_projects,
    get_db,
    get_contact_by_slug,
    get_graph_data,
    get_project_by_slug,
    get_report,
    get_reports,
    get_stats,
    search_contacts,
    search_projects,
)
from core.report_agent import generate_agent_report, html_to_pdf
from models.project import ContactsResponse, Project, ProjectsResponse, StatsResponse

router = APIRouter()


def _project_slug(doc: dict) -> str:
    """Extract slug from project document."""
    code = str(doc.get("code") or "").strip()
    if code:
        return code
    return str(doc.get("entity_key") or doc.get("_id") or doc.get("name") or "").strip()


def _normalize_developers(doc: dict) -> list[str]:
    """Normalize developer field to list of strings."""
    value = doc.get("developer") or []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str) and value.strip():
        return [value.strip()]
    return []


def _value_str(doc: dict) -> str:
    """Format project value to display string."""
    value = doc.get("value_billion_vnd")
    if value is None or value == "":
        return str(doc.get("value_str") or "")
    try:
        number = float(value)
    except Exception:
        return str(doc.get("value_str") or value)
    return f"{number:,.3f} tỷ VND".replace(",", "_").replace(".", ",").replace("_", ".")


def _project_to_frontend(doc: dict) -> dict:
    """Convert MongoDB project document to frontend format."""
    return {
        "slug": _project_slug(doc),
        "name": str(doc.get("name") or ""),
        "province": str(doc.get("province") or ""),
        "district": str(doc.get("district") or ""),
        "address": str(doc.get("address") or ""),
        "developer": _normalize_developers(doc),
        "value_billion_vnd": doc.get("value_billion_vnd"),
        "value_str": _value_str(doc),
        "status": str(doc.get("status") or ""),
        "phase": str(doc.get("design_stage") or doc.get("phase") or ""),
        "dev_type": str(doc.get("build_type") or ""),
        "owner_type": str(doc.get("category") or ""),
        "site_area": str(doc.get("area_sqm") or doc.get("site_area") or ""),
        "floor_area": str(doc.get("floor_area") or ""),
        "floors": str(doc.get("floors") or ""),
        "groundbreaking": str(doc.get("start_date") or ""),
        "handover": str(doc.get("end_date") or ""),
        "type_tags": [str(doc.get("category") or "").strip()] if doc.get("category") else [],
        "notes": str(doc.get("description") or doc.get("notes") or ""),
        "source_file": str(doc.get("source_file") or ""),
    }


def _contact_to_frontend(doc: dict) -> dict:
    """Convert MongoDB contact document to frontend format."""
    return {
        "slug": str(doc.get("entity_key") or doc.get("_id") or doc.get("full_name") or "").strip(),
        "name": str(doc.get("full_name") or doc.get("name") or ""),
        "company": str(doc.get("company") or ""),
        "role": str(doc.get("role") or ""),
        "phone": str(doc.get("phone") or ""),
        "email": str(doc.get("email") or ""),
        "address": str(doc.get("address") or ""),
        "image_path": str(doc.get("image_path") or ""),
    }


def _project_filters_from_request(
    province: str | None,
    developer: str | None,
    status: str | None,
) -> dict[str, str]:
    """Build filter dict from request params."""
    filters: dict[str, str] = {}
    if province:
        filters["province"] = province
    if developer:
        filters["company"] = developer
    if status:
        filters["status"] = status
    return filters


def _row_identity_filter(source: str, row_id: str) -> dict:
    """Build resilient identity filter for table row update."""
    clauses = [{"entity_key": row_id}, {"_id": row_id}]
    if source == "projects":
        clauses.insert(1, {"code": row_id})
        clauses.append({"name": row_id})
    else:
        clauses.append({"full_name": row_id})

    try:
        clauses.append({"_id": ObjectId(row_id)})
    except Exception:
        pass

    return {"$or": clauses}


def _project_update_doc(values: dict[str, str]) -> dict:
    """Map editable project fields from table payload to Mongo document fields."""
    out: dict = {}
    for key, raw_value in values.items():
        value = "" if raw_value is None else str(raw_value).strip()

        if key in {"name", "province", "district", "address", "status", "floors", "notes", "value_str"}:
            target = "description" if key == "notes" else key
            out[target] = value
        elif key == "developer":
            out["developer"] = [x.strip() for x in value.split(",") if x.strip()]
        elif key == "type_tags":
            tags = [x.strip() for x in value.split(",") if x.strip()]
            out["category"] = tags[0] if tags else ""
        elif key == "phase":
            out["design_stage"] = value
        elif key == "site_area":
            out["site_area"] = value
            out["area_sqm"] = value
        elif key == "groundbreaking":
            out["start_date"] = value
        elif key == "handover":
            out["end_date"] = value

    return out


def _project_insert_doc(values: dict[str, str]) -> dict:
    """Build a project document for insert with permissive defaults."""
    doc: dict = {
        "name": "",
        "province": "",
        "district": "",
        "address": "",
        "status": "",
        "floors": "",
        "description": "",
        "value_str": "",
        "design_stage": "",
        "site_area": "",
        "area_sqm": "",
        "start_date": "",
        "end_date": "",
        "category": "",
        "developer": [],
    }
    doc.update(_project_update_doc(values))
    return doc


def _contact_update_doc(values: dict[str, str]) -> dict:
    """Map editable contact fields from table payload to Mongo document fields."""
    out: dict = {}
    for key, raw_value in values.items():
        value = "" if raw_value is None else str(raw_value).strip()

        if key == "name":
            out["full_name"] = value
        elif key in {"company", "role", "phone", "email", "address"}:
            out[key] = value

    return out


def _contact_insert_doc(values: dict[str, str]) -> dict:
    """Build a contact document for insert with permissive defaults."""
    doc: dict = {
        "full_name": "",
        "company": "",
        "role": "",
        "phone": "",
        "email": "",
        "address": "",
        "image_path": "",
    }
    doc.update(_contact_update_doc(values))
    return doc


@router.get("/stats", response_model=StatsResponse)
async def stats():
    """Get comprehensive database statistics."""
    raw = await get_stats()
    counts = raw.get("counts", {})
    provinces = [
        item.get("province", "")
        for item in raw.get("projects_by_province", [])
        if item.get("province")
    ]
    if not provinces:
        provinces = []
    statuses = [
        item.get("status", "")
        for item in raw.get("project_value_by_status", [])
        if item.get("status")
    ]
    if not statuses:
        statuses = []
    return {
        "total_projects": int(counts.get("projects", 0)),
        "total_contacts": int(counts.get("contacts", 0)),
        "total_relationships": int(counts.get("relationships", 0)),
        "total_value_billion_vnd": float(raw.get("project_value_total_billion_vnd", 0) or 0),
        "provinces": provinces,
        "statuses": statuses,
    }


@router.get("/projects", response_model=ProjectsResponse)
async def list_projects(
    province: str = Query(None),
    developer: str = Query(None),
    status: str = Query(None),
    type_tag: str = Query(None),
    search: str = Query(None),
    limit: int = Query(100, le=500),
    page: int = Query(1, ge=1),
):
    """List projects with pagination and filters."""
    skip = (page - 1) * limit
    filters = _project_filters_from_request(province, developer, status)
    if type_tag:
        filters["category"] = type_tag
    items = await search_projects(query=search, filters=filters, skip=skip, limit=limit)
    total = await count_projects(query=search, filters=filters)
    data = [_project_to_frontend(item) for item in items]
    return {"data": data, "total": total, "page": page, "limit": limit}


@router.get("/projects/{slug}")
async def get_one_project(slug: str):
    """Get a single project by slug."""
    project = await get_project_by_slug(slug)
    if project:
        return _project_to_frontend(project)
    raise HTTPException(404, "Project not found")


@router.get("/contacts", response_model=ContactsResponse)
async def list_contacts(
    company: str = Query(None),
    search: str = Query(None),
    limit: int = Query(100, le=500),
    page: int = Query(1, ge=1),
):
    """List contacts with pagination and filters."""
    skip = (page - 1) * limit
    filters = {"company": company} if company else {}
    items = await search_contacts(query=search, filters=filters, skip=skip, limit=limit)
    total = await count_contacts(query=search, filters=filters)
    data = [_contact_to_frontend(item) for item in items]
    return {"data": data, "total": total}


@router.get("/contacts/{slug}")
async def get_one_contact(slug: str):
    """Get a single contact by slug."""
    contact = await get_contact_by_slug(slug)
    if contact:
        return _contact_to_frontend(contact)
    raise HTTPException(404, "Contact not found")


@router.get("/graph")
async def graph_data():
    """Get graph data (nodes and links) for visualization, enriched with GraphRAG community data."""
    from core.graphrag_client import get_graphrag_client

    raw = await get_graph_data()
    nodes: list[dict] = []

    # Get community map from GraphRAG (best-effort — may be empty if index not built yet)
    try:
        rag = get_graphrag_client()
        if rag._built:
            community_map = rag.get_entity_community_map()
        else:
            community_map = {}
    except Exception:
        community_map = {}

    for node in raw.get("nodes", []):
        node_type = str(node.get("type") or "").strip() or "project"
        color = {
            "project": "#6366F1",
            "company": "#F59E0B",
            "contact": "#10B981",
            "person": "#10B981",
        }.get(node_type, "#6366F1")

        node_id = str(node.get("id") or "")
        mapped = {
            "id": node_id,
            "name": str(node.get("name") or ""),
            "type": "person" if node_type == "contact" else node_type,
            "color": color,
            "size": 6 if node_type == "project" else (5 if node_type in {"company", "contact", "person"} else 4),
        }

        if node.get("province") is not None:
            mapped["province"] = str(node.get("province") or "")
        if node_type == "project" and node.get("value_billion_vnd") is not None:
            mapped["value_str"] = _value_str(node)
        if node_type in {"contact", "person"}:
            mapped["hidden"] = node.get("hidden", False)
        if node.get("hidden"):
            mapped["hidden"] = True

        # Attach community_id for graph coloring
        if node_id in community_map:
            mapped["community_id"] = community_map[node_id]

        nodes.append(mapped)

    links = []
    for link in raw.get("links", []):
        links.append(
            {
                "source": str(link.get("source") or ""),
                "target": str(link.get("target") or ""),
                "type": str(link.get("type") or "related"),
            }
        )

    return {"nodes": nodes, "links": links}


@router.get("/graph/communities")
async def graph_communities():
    """Return community summaries for graph visualization overlay."""
    from core.graphrag_client import get_graphrag_client
    rag = get_graphrag_client()
    # Trigger index build if not done yet
    await rag.ensure_index()
    return {"communities": rag.get_communities_summary()}


@router.get("/export/excel")
async def export_excel(
    province: str = Query(None),
    developer: str = Query(None),
    status: str = Query(None),
    search: str = Query(None),
):
    """Export filtered projects to Excel."""
    filters = _project_filters_from_request(province, developer, status)
    items = await search_projects(query=search, filters=filters, skip=0, limit=1000)
    data = [_project_to_frontend(item) for item in items]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dự án LeadsMap"

    header_fill = PatternFill("solid", fgColor="1E1E2E")
    header_font = Font(bold=True, color="E2E8F0", size=11)
    headers = [
        "Slug", "Tên dự án", "Tỉnh", "Quận/Huyện", "Địa chỉ",
        "Chủ đầu tư", "Giá trị", "Trạng thái", "Giai đoạn",
        "Loại hình", "Diện tích đất", "Số tầng", "Khởi công", "Bàn giao",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(data, 2):
        devs = p.get("developer", [])
        devs_str = ", ".join(devs) if isinstance(devs, list) else str(devs)
        tags = p.get("type_tags", [])
        tags_str = ", ".join(tags) if isinstance(tags, list) else str(tags)

        ws.cell(row=row_idx, column=1, value=p.get("slug", ""))
        ws.cell(row=row_idx, column=2, value=p.get("name", ""))
        ws.cell(row=row_idx, column=3, value=p.get("province", ""))
        ws.cell(row=row_idx, column=4, value=p.get("district", ""))
        ws.cell(row=row_idx, column=5, value=p.get("address", ""))
        ws.cell(row=row_idx, column=6, value=devs_str)
        ws.cell(row=row_idx, column=7, value=p.get("value_str", ""))
        ws.cell(row=row_idx, column=8, value=p.get("status", ""))
        ws.cell(row=row_idx, column=9, value=p.get("phase", ""))
        ws.cell(row=row_idx, column=10, value=tags_str)
        ws.cell(row=row_idx, column=11, value=p.get("site_area", ""))
        ws.cell(row=row_idx, column=12, value=p.get("floors", ""))
        ws.cell(row=row_idx, column=13, value=p.get("groundbreaking", ""))
        ws.cell(row=row_idx, column=14, value=p.get("handover", ""))

        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="F8FAFC")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leadsmap_export.xlsx"},
    )


@router.post("/export/table")
def export_table_from_chat(payload: dict = Body(...)):
    """Export custom table data to Excel."""
    headers: list[str] = payload.get("headers", [])
    rows: list[list] = payload.get("rows", [])
    title: str = payload.get("title", "LeadsMap Export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]

    hdr_fill = PatternFill("solid", fgColor="1E1E2E")
    hdr_font = Font(bold=True, color="E2E8F0", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx % 2 == 0:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor="F8FAFC")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = title.replace(" ", "_")[:30]

    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{safe_title}_{ts}.xlsx"
        file_path = Path(__file__).resolve().parents[1] / "reports" / file_name
        file_path.parent.mkdir(parents=True, exist_ok=True)
        with open(str(file_path), "wb") as f:
            f.write(buf.getvalue())
    except Exception:
        pass

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_title}.xlsx"},
    )


@router.get("/fields")
def available_fields():
    """Get available fields for projects and contacts."""
    return {
        "projects": [
            {"key": "name", "label": "Tên dự án"},
            {"key": "province", "label": "Tỉnh/TP"},
            {"key": "district", "label": "Quận/Huyện"},
            {"key": "address", "label": "Địa chỉ"},
            {"key": "developer", "label": "Chủ đầu tư"},
            {"key": "value_str", "label": "Giá trị"},
            {"key": "status", "label": "Trạng thái"},
            {"key": "phase", "label": "Giai đoạn"},
            {"key": "type_tags", "label": "Loại hình"},
            {"key": "site_area", "label": "Diện tích đất"},
            {"key": "floors", "label": "Số tầng"},
            {"key": "groundbreaking", "label": "Khởi công"},
            {"key": "handover", "label": "Bàn giao"},
            {"key": "notes", "label": "Ghi chú"},
        ],
        "contacts": [
            {"key": "name", "label": "Họ tên"},
            {"key": "company", "label": "Công ty"},
            {"key": "role", "label": "Chức vụ"},
            {"key": "phone", "label": "Số điện thoại"},
            {"key": "email", "label": "Email"},
            {"key": "address", "label": "Địa chỉ"},
        ],
    }


@router.post("/query")
async def flexible_query(payload: dict = Body(...)):
    """Execute flexible query with column selection."""
    source = payload.get("source", "projects")
    columns: list[str] = payload.get("columns", [])
    filters: dict = payload.get("filters", {})
    try:
        raw_limit = int(payload.get("limit", 100))
    except Exception:
        raw_limit = 100
    try:
        raw_page = int(payload.get("page", 1))
    except Exception:
        raw_page = 1

    limit = min(max(raw_limit, 1), 1000)
    page = max(raw_page, 1)
    skip = (page - 1) * limit
    search = filters.get("search")

    if source == "contacts":
        company = filters.get("company") or filters.get("developer")
        contact_filters: dict[str, str] = {}
        if company:
            contact_filters["company"] = company

        items = await search_contacts(
            query=search,
            filters=contact_filters,
            skip=skip,
            limit=limit,
        )
        total = await count_contacts(query=search, filters=contact_filters)
        data = [_contact_to_frontend(item) for item in items]
    else:
        province = filters.get("province")
        developer = filters.get("developer") or filters.get("company")
        status = filters.get("status")
        items = await search_projects(
            query=search,
            filters=_project_filters_from_request(province, developer, status),
            skip=skip,
            limit=limit,
        )
        total = await count_projects(
            query=search,
            filters=_project_filters_from_request(province, developer, status),
        )
        data = [_project_to_frontend(item) for item in items]

    if columns:
        result = []
        for row in data:
            filtered: dict = {}
            for col in columns:
                v = row.get(col, "")
                if isinstance(v, list):
                    v = ", ".join(str(x) for x in v)
                elif v is None:
                    v = ""
                filtered[col] = v
            filtered["__row_id"] = str(row.get("slug") or "")
            result.append(filtered)
        data = result

    return {"data": data, "total": total, "page": page, "limit": limit}


@router.post("/query/save")
async def save_flexible_query(payload: dict = Body(...)):
    """Persist edited table rows back to MongoDB."""
    source = str(payload.get("source") or "projects")
    changes = payload.get("changes") or []
    deletes = payload.get("deletes") or []

    if source not in {"projects", "contacts"}:
        raise HTTPException(400, "Invalid source")
    if not isinstance(changes, list):
        raise HTTPException(400, "changes must be an array")
    if not isinstance(deletes, list):
        raise HTTPException(400, "deletes must be an array")

    db = await get_db()
    collection = "projects" if source == "projects" else "contacts"

    updated = 0
    inserted = 0
    deleted = 0
    skipped = 0
    errors: list[str] = []

    for raw_row_id in deletes:
        row_id = str(raw_row_id or "").strip()
        if not row_id:
            skipped += 1
            continue
        try:
            result = await db[collection].delete_one(_row_identity_filter(source, row_id))
            if result.deleted_count > 0:
                deleted += 1
            else:
                skipped += 1
        except Exception as e:
            errors.append(f"delete:{row_id}: {e}")

    for idx, change in enumerate(changes):
        change_obj = change if isinstance(change, dict) else {}
        row_id = str(change_obj.get("row_id") or "").strip()
        values = change_obj.get("values") or {}

        if not isinstance(values, dict):
            skipped += 1
            continue

        normalized_values = {
            str(key): "" if raw_value is None else str(raw_value)
            for key, raw_value in values.items()
        }

        try:
            if row_id:
                update_doc = _project_update_doc(normalized_values) if source == "projects" else _contact_update_doc(normalized_values)
                if not update_doc:
                    skipped += 1
                    continue

                update_doc["updated_at"] = datetime.utcnow()
                result = await db[collection].update_one(
                    _row_identity_filter(source, row_id),
                    {"$set": update_doc},
                )
                if result.matched_count > 0:
                    updated += 1
                else:
                    skipped += 1
            else:
                now = datetime.utcnow()
                insert_doc = _project_insert_doc(normalized_values) if source == "projects" else _contact_insert_doc(normalized_values)
                insert_doc["created_at"] = now
                insert_doc["updated_at"] = now
                await db[collection].insert_one(insert_doc)
                inserted += 1
        except Exception as e:
            row_label = row_id or f"new_row_{idx + 1}"
            errors.append(f"{row_label}: {e}")

    return {
        "success": len(errors) == 0,
        "updated": updated,
        "inserted": inserted,
        "deleted": deleted,
        "skipped": skipped,
        "errors": errors[:10],
    }


def _reports_root() -> Path:
    return Path(__file__).resolve().parents[1] / "reports"


def _safe_report_path(raw_path: str) -> Path:
    root = _reports_root().resolve()
    path = Path(raw_path)
    if not path.is_absolute():
        path = root / path
    resolved = path.resolve()
    if not str(resolved).lower().startswith(str(root).lower()):
        raise HTTPException(403, "Invalid report path")
    return resolved


def _report_to_frontend(doc: dict) -> dict:
    file_path = str(doc.get("file_path") or "")
    created_at = str(doc.get("created_at") or doc.get("timestamp") or "")
    suffix = Path(file_path).suffix.lower()
    return {
        "id": str(doc.get("id") or ""),
        "title": str(doc.get("title") or "Báo cáo"),
        "file_path": file_path,
        "query": str(doc.get("query") or ""),
        "timestamp": created_at,
        "kind": "html" if suffix == ".html" else ("pdf" if suffix == ".pdf" else "excel"),
        "html_available": suffix == ".html",
    }


@router.get("/reports")
async def reports():
    """List saved reports."""
    items = await get_reports()
    return {"data": [_report_to_frontend(item) for item in items]}


@router.post("/reports/agent")
async def create_agent_report(payload: dict = Body(...)):
    """Generate a Vietnamese Agent report with HTML preview and PDF export."""
    query = str(payload.get("query") or "").strip()
    language = str(payload.get("language") or "vi").strip() or "vi"
    report = await generate_agent_report(query=query, language=language)
    return {"success": True, "report": report}


@router.get("/reports/{report_id}/html")
async def report_html(report_id: str):
    """Return stored HTML for preview."""
    report = await get_report(report_id)
    if not report:
        raise HTTPException(404, "Report not found")
    path = _safe_report_path(str(report.get("file_path") or ""))
    if path.suffix.lower() != ".html" or not path.exists():
        raise HTTPException(404, "HTML preview is not available")
    return {"html": path.read_text(encoding="utf-8"), "report": _report_to_frontend(report)}


@router.get("/reports/{report_id}/download")
async def report_download(report_id: str):
    """Download a saved report. HTML Agent reports are exported as PDF."""
    report = await get_report(report_id)
    if not report:
        raise HTTPException(404, "Report not found")
    path = _safe_report_path(str(report.get("file_path") or ""))
    if not path.exists():
        raise HTTPException(404, "Report file not found")

    download_path = path
    media_type = "application/octet-stream"
    if path.suffix.lower() == ".html":
        pdf_path = path.with_suffix(".pdf")
        if not pdf_path.exists():
            html_text = path.read_text(encoding="utf-8")
            html_to_pdf(html_text, pdf_path)
        download_path = pdf_path
        media_type = "application/pdf"
    elif path.suffix.lower() == ".pdf":
        media_type = "application/pdf"
    elif path.suffix.lower() == ".xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return FileResponse(
        str(download_path),
        media_type=media_type,
        filename=download_path.name,
    )


@router.post("/projects/{slug}/summary")
async def project_summary(slug: str):
    """Get project summary."""
    project = await get_project_by_slug(slug)
    if not project:
        raise HTTPException(404, "Project not found")

    mapped = _project_to_frontend(project)
    return {
        "summary": f"Tổng quan dự án {mapped['name']}",
        "report_id": None,
        "project": mapped["name"],
    }
