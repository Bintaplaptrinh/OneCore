# -*- coding: utf-8 -*-
"""
Script bổ sung dữ liệu thiếu cho TotalParking_Danh sach du an_260315.xlsx
- Chỉ điền vào các ô TRỐNG, không ghi đè dữ liệu có sẵn
- Nguồn: tìm kiếm internet + suy luận từ tên dự án
"""

from openpyxl import load_workbook
import pandas as pd

INPUT_FILE  = 'D:/HaiVo Leads/1_project/TotalParking_Danh sach du an_260315.xlsx'
OUTPUT_FILE = 'D:/HaiVo Leads/1_project/TotalParking_Danh sach du an_260315_updated.xlsx'

wb = load_workbook(INPUT_FILE)
ws = wb.active

# ── Build column index map (header row = row 1) ────────────────────────────
headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

def set_if_empty(excel_row, field, value):
    col = headers.get(field)
    if not col:
        return
    cell = ws.cell(excel_row, col)
    if cell.value is None or str(cell.value).strip() in ('', 'nan', 'NaN'):
        cell.value = value

# excel_row = df_index + 2  (row 1 = header)

# ═══════════════════════════════════════════════════════════════════════════
# 1. Prj_Client — dữ liệu từ tìm kiếm internet
# ═══════════════════════════════════════════════════════════════════════════
client_updates = {
    0:  'Đại học Sư phạm Đà Nẵng',           # DH_SPDN (from description)
    1:  'Perdana ParkCity (VIDC)',             # ParkCity LTT Hanoi
    2:  'Đầu tư Đại La',                      # CC_800 DVC
    6:  'Kim Oanh Group',                      # Hoà Lân Smart City
    12: 'Novaland',                            # Nova
    19: 'Tập đoàn Khai Sơn',                  # CC_KhaiSon_LB
    23: 'Wealthcons',                          # GARRYA_DN
    30: 'Rita Võ',                             # RITA VÕ QUẬN 5
}

for idx, client in client_updates.items():
    set_if_empty(idx + 2, 'Prj_Client', client)

# ═══════════════════════════════════════════════════════════════════════════
# 2. Cons_Type — suy luận từ tên dự án / loại công trình
# ═══════════════════════════════════════════════════════════════════════════
cons_type_updates = {
    0:   'PARKING BULDING',   # DH_SPDN - nhà đỗ xe cao tầng ĐH Sư phạm
    1:   'Complex Building',  # ParkCity - khu đô thị hỗn hợp 77ha
    2:   'Apartment',         # CC_800 DVC - chung cư
    3:   'Office',            # Building 511A NDT - tòa nhà văn phòng
    4:   'Apartment',         # 02 NHC - nhà ở
    5:   'Office',            # VP hạng A - văn phòng
    6:   'Complex Building',  # Hoà Lân Smart City - khu đô thị
    7:   'Others',            # Nhà máy KCN - khu công nghiệp
    8:   'Apartment',         # Sông Cẩm Lệ - dự án căn hộ ven sông
    9:   'Complex Building',  # TTTM Bitis - trung tâm thương mại
    11:  'Apartment',         # Sát 02 Nguyễn Lộ Trạch - căn hộ
    12:  'Complex Building',  # Nova (Novaland) - khu phức hợp
    14:  'Apartment',         # CC Nguyen Viet Hong - chung cư
    17:  'Complex Building',  # MinhQuan_Com,.PY - phức hợp thương mại
    18:  'Complex Building',  # CAND_N02 - tổ hợp CAND
    21:  'Apartment',         # CC DT743 - chung cư
    22:  'Office',            # TT ĐT CNTT_PHÚ DIỄN_HN - trung tâm đào tạo
    23:  'Hotel',             # GARRYA_DN - khách sạn 5 sao
    25:  'Office',            # Marina IFC - trung tâm tài chính
    27:  'Complex Building',  # One Era_EVC - khu phức hợp
    28:  'Others',            # Car Lift _Thảo Điền Q2 - sản phẩm nâng xe
    29:  'Hotel',             # 01 NGÔ MÂY - khách sạn
    30:  'Apartment',         # RITA VÕ QUẬN 5 - căn hộ
    31:  'Others',            # Turntable Bike - mâm xoay xe máy
    32:  'Office',            # HFIC Q1 - văn phòng tài chính
    33:  'Others',            # Trạm Sạc D9 - trạm sạc EV
    34:  'PARKING BULDING',   # Hẻm 51HB-D11 - đầu tư bãi đỗ xe tư nhân
    35:  'Complex Building',  # GEM DN - khu giải trí tổng hợp
    36:  'Complex Building',  # THD-Q1 - tổ hợp SAMCO
    37:  'Others',            # Scissor Lift - sản phẩm nâng kéo
    38:  'Apartment',         # HH01 Trung Van_HaNoi - chung cư
    39:  'Apartment',         # 1283 Giai Phong - căn hộ
    40:  'Complex Building',  # THE GLOBAL CITY - khu đô thị 117ha
    41:  'Office',            # Trung Tam Doi Moi Sang Tao_DN - trung tâm đổi mới
    42:  'PARKING BULDING',   # Tower Parking_HN - tháp đỗ xe
    43:  'Complex Building',  # VEGA ĐAN PHƯỢNG - khu đô thị KDI
    44:  'Complex Building',  # Vo Van Kiet (KITA) - khu phức hợp
    45:  'Others',            # Barrier Chống Ngập - sản phẩm ngăn lũ
    46:  'Apartment',         # 42 Chu Manh Trinh - căn hộ
    47:  'Apartment',         # NOXH Lạc Hồng - nhà ở xã hội
    48:  'Apartment',         # 151 Tran Quang Khai - căn hộ
    49:  'Office',            # Nguyen Van Thu - tòa nhà VP
    50:  'Complex Building',  # 180 Pallets (Gamuda) - thương mại/logistics
    52:  'Hotel',             # Khách sạn Nguyễn Kim Huế
    53:  'Office',            # Tòa nhà Trần Cao Vân_HCM
    54:  'PARKING BULDING',   # Bãi xe_HN - BV Bạch Mai parking
    55:  'Apartment',         # CT1-CT2_DaNang - chung thư cư
    56:  'Apartment',         # F3-Thanh Liệt-HN - căn hộ
    57:  'Apartment',         # IA25 - căn hộ Thái Nam Land
    58:  'Complex Building',  # Nam Cấm_Nghệ An - khu phức hợp
    59:  'Apartment',         # Handico6 - chung cư Handico
    60:  'Apartment',         # Long Xuyên - dự án căn hộ
    61:  'Apartment',         # Long An - dự án căn hộ
    62:  'Complex Building',  # Tân Cảng - khu cảng thương mại
    63:  'Office',            # Khu VP CẢNG CÁT LÁI - văn phòng cảng
    64:  'Complex Building',  # Tân Thuận - khu phức hợp
    65:  'Apartment',         # A11 - căn hộ AZ Thăng Long
    66:  'Complex Building',  # Genco3 - tổ hợp điện lực
    67:  'Apartment',         # The Emerald 68 - chung cư Lê Phong
    68:  'Apartment',         # E2 Yên Hòa - chung cư
    69:  'Apartment',         # SORA 3 (C16) - căn hộ
    70:  'Office',            # MP13-Kiểm định - văn phòng/kiểm định
    71:  'Office',            # Văn Phòng_23 DBP - văn phòng tư nhân
    72:  'Complex Building',  # Gladia - biệt thự + shophouse Keppel+Khang Điền
    73:  'Office',            # Mitalab Đà nẵng - phòng lab/văn phòng
    74:  'Hotel',             # Khach San Long Xuyen
    75:  'Office',            # VP 38 BÍCH CÂU - văn phòng
    78:  'Office',            # VP THẢO ĐIỀN - văn phòng
    79:  'Office',            # TÒA NHÀ 204 HẢI PHÒNG
    80:  'Hotel',             # KS HANG THAN - khách sạn
    81:  'Hotel',             # DA NANG SILK TOWER 1 - khách sạn 4 sao
    82:  'Office',            # TÒA NHÀ VP QUẬN 1
    83:  'Complex Building',  # THE SPIRIT OF SAIGON - phức hợp VP+Hotel+CC
    84:  'Office',            # OFFICE 2F NGUYEN THANH Y
    85:  'Apartment',         # ROXANA_BINH DUONG - chung cư Tường Phong
    87:  'Hotel',             # KS NGÔ MÂY - khách sạn
    88:  'Complex Building',  # GENCO 3_HCM - tổ hợp điện lực
    89:  'Apartment',         # SETIA RESIDENCE_BÌNH DƯƠNG - SP Setia
    90:  'Complex Building',  # CT6 TIMES SQUARE - phức hợp thương mại
    91:  'Complex Building',  # SYCAMORE B14 - khu đô thị CapitaLand
    93:  'Apartment',         # EATON PARK - căn hộ Gamuda Land
    94:  'PARKING BULDING',   # BÃI ĐẬU XE HT01-05_LONG AN
    96:  'Apartment',         # CARMEL - căn hộ
    97:  'Hotel',             # ELLIE HOTEL (ĐÀ NẴNG)
    98:  'Complex Building',  # SYCAMORE B8 - khu đô thị CapitaLand
    99:  'Apartment',         # H5H7 - chung cư
    100: 'Apartment',         # LUMI_HN - căn hộ CapitaLand
    101: 'Complex Building',  # NEXUS 3 - tổ hợp
    102: 'Hotel',             # SAIGON-DANANG - Saigon Tourist hotel
    103: 'Apartment',         # F12 - căn hộ
    104: 'Office',            # APG BUILDING - tòa nhà văn phòng
    105: 'Apartment',         # LACASA (CẢI TẠO) - chung cư cải tạo
    106: 'Office',            # TRUONG DINH TOWER - văn phòng + TMDV
    107: 'PARKING BULDING',   # SKY GARAGE - bãi xe
    108: 'Apartment',         # POLARIS - chung cư Bcons
    109: 'Apartment',         # VINACONEX - chung cư Vinaconex
    110: 'Office',            # SIB ĐÀ NẴNG - văn phòng
    111: 'PARKING BULDING',   # CÔNG VIÊN ĐỖ XE - bãi đỗ xe công cộng
    112: 'Complex Building',  # BIỆT THỰ ĐÀ NẴNG - biệt thự cao cấp
    113: 'Complex Building',  # MASTERISE THẢO ĐIỀN (BẢO TRÌ) - bảo trì hệ thống
    114: 'Apartment',         # LUX 1 - PHASE 2 (CAMBODIA) - căn hộ Campuchia
    115: 'PARKING BULDING',   # FULL AUTOMATION SYSTEM - bãi xe tự động
    116: 'PARKING BULDING',   # BẾN XE MIỀN ĐÔNG - bến xe mới
}

for idx, cons_type in cons_type_updates.items():
    set_if_empty(idx + 2, 'Cons_Type', cons_type)

# ═══════════════════════════════════════════════════════════════════════════
# 3. Cons_Phase — suy luận từ Prj_Progress
# ═══════════════════════════════════════════════════════════════════════════
progress_to_phase = {
    '1 Tiềm năng💡':          'Ý tưởng',
    '2 Tư vấn kỹ thuật🛠️':   'Thiết kế',
    '3 Chào giá🚀':           'Thiết kế',
    '4 Theo dõi🚦':           'Thi công Xây dựng',
    '5 Win🏁':                 'Thi công Xây dựng',
    '6 Lost🚧':               'Dự án đã hoàn thành',
}

df = pd.read_excel(INPUT_FILE)
for idx, row in df.iterrows():
    progress = str(row.get('Prj_Progress', '')).strip()
    phase = progress_to_phase.get(progress)
    if phase:
        set_if_empty(idx + 2, 'Cons_Phase', phase)

# Override Cons_Phase cho các dự án đã xác nhận qua internet
confirmed_phase = {
    23:  'Thi công Xây dựng',  # GARRYA_DN - đang thi công tầng 7 (9/2024)
    81:  'Thi công Xây dựng',  # DA NANG SILK TOWER 1 - đang thi công
    93:  'Thi công Xây dựng',  # EATON PARK - đã bàn giao 90%
    100: 'Thi công Xây dựng',  # LUMI_HN - đang thi công
    106: 'Dự án đã hoàn thành', # TRUONG DINH TOWER - hoàn thành Q1/2025
}
for idx, phase in confirmed_phase.items():
    set_if_empty(idx + 2, 'Cons_Phase', phase)

# ═══════════════════════════════════════════════════════════════════════════
# 4. Sol_Category — chỉ điền khi rõ ràng từ tên dự án / loại sản phẩm
# ═══════════════════════════════════════════════════════════════════════════
sol_cat_updates = {
    27:  'EVC',                    # One Era_EVC - sạc xe điện
    28:  'Car Lift',               # Car Lift _Thảo Điền Q2
    31:  'Turntable Motorbike',    # Turntable Bike - mâm xoay xe máy
    33:  'EVC',                    # Trạm Sạc D9 - trạm sạc EV
    34:  'FullAutomation',         # Hẻm 51HB-D11 - đầu tư bãi xe tự động Q11
    42:  'FullAutomation',         # Tower Parking_HN - tháp đỗ xe tự động
    45:  'Flood Barrier',          # Barrier Chống Ngập
    54:  'Puzzle',                 # Bãi xe_HN - BV Bạch Mai (thường dùng puzzle)
    94:  'FullAutomation',         # BÃI ĐẬU XE HT01-05_LONG AN - bãi xe chuyên dụng
    107: 'FullAutomation',         # SKY GARAGE
    111: 'FullAutomation',         # CÔNG VIÊN ĐỖ XE
    115: 'FullAutomation',         # FULL AUTOMATION SYSTEM
    116: 'FullAutomation',         # BẾN XE MIỀN ĐÔNG
}

for idx, sol_cat in sol_cat_updates.items():
    set_if_empty(idx + 2, 'Sol_Category', sol_cat)

# ═══════════════════════════════════════════════════════════════════════════
# 5. Prj_Description — viết mô tả cho các dự án lớn tìm được thông tin
# ═══════════════════════════════════════════════════════════════════════════
descriptions = {
    1: ("Khu đô thị ParkCity Hà Nội tại ngã tư Tố Hữu - Lê Trọng Tấn, quận Hà Đông. "
        "Chủ đầu tư: Perdana ParkCity (Malaysia) / VIDC. Tổng diện tích 77 ha, bao gồm "
        "biệt thự, căn hộ cao cấp (7.000 căn), trung tâm thương mại The Linc và 13,7 ha "
        "cây xanh. Dự án đang trong giai đoạn phát triển các tiểu khu mới."),

    6: ("Khu đô thị Hòa Lân (One World / The One World) tại phường Thuận Giao, TP Thuận An, "
        "Bình Dương. Chủ đầu tư: Kim Oanh Group (Việt-Nhật). Diện tích gần 50 ha, vốn đầu "
        "tư 1 tỷ USD. Sản phẩm: shophouse, nhà phố liền kề, biệt thự compound, căn hộ. "
        "Dự án đang trong giai đoạn thiết kế và phát triển phân khu."),

    19: ("Chung cư Khai Sơn City Long Biên, Hà Nội. Chủ đầu tư: Tập đoàn Khai Sơn. "
         "Khu đô thị rộng 180 ha thuộc địa phận Ngọc Thụy, Thượng Thanh, Đức Giang, quận Long Biên. "
         "Quy mô 6 tòa cao 21 tầng, 3 tầng hầm, tổng 1.100 căn hộ diện tích 77-120 m². "
         "Dự án có hồ điều hòa 22 ha, vị trí đẹp giữa sông Hồng và sông Đuống."),

    23: ("Dự án khách sạn GARRYA Đà Nẵng (thương hiệu Banyan Tree Group - Singapore) tại trung "
         "tâm thành phố Đà Nẵng. Chủ đầu tư: Wealthcons. Quy mô 1 tòa 9 tầng + 2 tầng hầm, "
         "GFA 15.503 m², tiêu chuẩn 5 sao. Tiến độ: đang thi công (đã đổ sàn tầng 7, 9/2024)."),

    40: ("Khu đô thị The Global City tại quận 2 (TP. Thủ Đức), TP.HCM. Chủ đầu tư: "
         "Masterise Homes / SDI Corp. Quy mô 117,4 ha với hơn 10.000 căn hộ cao cấp, "
         "1.800 shophouse/SOHO, biệt thự và trung tâm thương mại Canal Walk. "
         "Đây là một trong những dự án BĐS lớn nhất TP.HCM."),

    43: ("Khu đô thị Vega City Đan Phượng, Hà Nội. Chủ đầu tư: KDI Holdings (liên doanh "
         "cựu Chủ tịch Sacombank và Phó Chủ tịch Techcombank). Dự án quy mô lớn tại huyện "
         "Đan Phượng, Hà Nội, kết hợp căn hộ, biệt thự, khu nghỉ dưỡng và thương mại. "
         "TotalParking đang chào giá hệ thống đỗ xe cho dự án."),

    67: ("Dự án The Emerald 68 tại 68 đường Vĩnh Phú 16, phường Vĩnh Phú, TP Thuận An, "
         "Bình Dương. Chủ đầu tư: Lê Phong Group. Quy mô 2 block 39 tầng + 3 tầng hầm, "
         "800 căn hộ. Nhà thầu chính: Coteccons. Đây là dự án hợp tác đầu tiên giữa "
         "Lê Phong Group và Coteccons Group."),

    72: ("Dự án Gladia by the Waters tại đường Võ Chí Công, Bình Trưng, TP.HCM. "
         "Chủ đầu tư: Keppel Land (Singapore) + Khang Điền. Tổng diện tích 11,8 ha, "
         "226 sản phẩm thấp tầng (nhà phố vườn, biệt thự song lập, biệt thự đơn lập, "
         "biệt thự tứ lập). Ra mắt chính thức 07/2025. Dự án đầu tiên tại Việt Nam đạt "
         "chứng chỉ BCA Green Mark District (Singapore)."),

    81: ("Dự án Đà Nẵng Silk Tower 1 (Tháp Lụa) tại đường Võ Nguyên Giáp, phường Bắc Mỹ An, "
         "quận Ngũ Hành Sơn, Đà Nẵng. Chủ đầu tư: Công ty TNHH Đầu Tư Tháp Lụa Đà Nẵng. "
         "Loại hình: khách sạn 4 sao + condotel. Diện tích đất 1.220 m², 2 tầng hầm + 25 tầng "
         "nổi, GFA 22.810 m², 193 căn nghỉ dưỡng. Đang trong giai đoạn thi công."),

    85: ("Dự án Roxana Plaza tại số 9/14, Đông Q., phường Vĩnh Phú, TP Thuận An, Bình Dương. "
         "Chủ đầu tư: Tường Phong. Quy mô 3 block 28 tầng, hơn 1.100 căn hộ + shophouse + "
         "officetel + penthouse. Dự án có lịch sử pháp lý phức tạp, hiện đang được xử lý "
         "để bàn giao cho khách hàng."),

    89: ("Dự án Setia Gardens Residences tại Lái Thiêu, TP Thuận An, Bình Dương. "
         "Chủ đầu tư: Setia Lái Thiêu (SP Setia Berhad - Malaysia, 50 năm kinh nghiệm). "
         "Tổng diện tích 12.389 m², 3 tòa tháp 27-28 tầng (Azalea, Begonia, Camellia), "
         "khoảng 865 căn hộ. Giá từ 42 triệu/m². Đây là dự án cao cấp chuẩn Malaysia tại VN."),

    93: ("Dự án Eaton Park tại đường Mai Chí Thọ, quận 2, TP.HCM. Chủ đầu tư: Gamuda Land "
         "(Malaysia). Quy mô 6 tòa cao 29-39 tầng, 1.968 căn hộ (1-3 phòng ngủ, penthouse) "
         "và 72 unit thương mại ven sông. Dự án bán 90% trong vài giờ khi mở bán (2025). "
         "Hiện đang thi công và bàn giao."),

    98: ("Khu đô thị Sycamore - CapitaLand tại TP Bình Dương (giao Hùng Vương - Võ Văn Kiệt). "
         "Chủ đầu tư: CapitaLand (Singapore) + Becamex IDC. Quy mô 59,32 ha gồm nhà thấp tầng, "
         "biệt thự, căn hộ (7 phân khu, 462 nhà thấp tầng + 3.300 căn hộ). "
         "Dự kiến bàn giao Q3/2025."),

    100: ("Dự án Lumi Hanoi tại Đại Lộ Thăng Long, Nam Từ Liêm, Hà Nội. Chủ đầu tư: "
          "CapitaLand Development (Singapore). Quy mô gần 5,6 ha, 9 tòa tháp cao 29-35 tầng, "
          "3.950 căn hộ cao cấp (1-3 phòng ngủ, duplex, penthouse, shophouse). Đang thi công."),

    106: ("Dự án Tòa nhà Văn phòng - TMDV Trường Định Tower, Hà Nội. Chủ đầu tư: Công ty CP "
          "Đầu tư Duy Trung Ý. Nhà thầu: SOL E&C. Loại hình: văn phòng + thương mại dịch vụ. "
          "Dự kiến hoàn thành và đưa vào sử dụng Q1/2025."),
}

for idx, desc in descriptions.items():
    set_if_empty(idx + 2, 'Prj_Description', desc)

# ═══════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT_FILE)
print(f"[OK] Da luu file: {OUTPUT_FILE}")

# ── Verify: show completion stats ──────────────────────────────────────────
df_out = pd.read_excel(OUTPUT_FILE)
fields = ['Prj_Client', 'Prj_Description', 'Sol_Category', 'Cons_Type', 'Cons_Phase']
print("\n-- Thong ke sau cap nhat (117 du an thuc, rows 0-116) --")
df_proj = df_out.iloc[:117]
for f in fields:
    filled = df_proj[f].notna().sum()
    pct = filled / len(df_proj) * 100
    print(f"  {f:25s}: {filled:3d}/117  ({pct:.0f}%)")
